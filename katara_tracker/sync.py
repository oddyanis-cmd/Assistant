"""Sync the edited tracking workbook back into a workbook + ODP deck.

Reading the workbook, it:
  * treats the per-row "Workflow Status" cell as the source of truth and
    re-files every member into the matching sheet,
  * reads new members from the "Create Slide" sheet,
  * rebuilds the Analysis totals,
  * regenerates the ODP deck (existing members keep their original slide with
    the status note updated; new members get a freshly cloned slide + photo),
    ordered Pending -> Approved -> Paid so status changes move slides.
"""

from __future__ import annotations

import os
import re
import sys

import openpyxl

from .excel_builder import COLUMNS, build_workbook
from .odp_parser import Client
from .theme import STATUSES

RATE_NUM_RE = re.compile(r"([\d][\d,\.]*)")


def _to_amount(value) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)
    m = RATE_NUM_RE.search(str(value).replace(" ", ""))
    return int(m.group(1).replace(",", "").split(".")[0]) if m else 0


def _header_index(ws, header_row: int) -> dict:
    out = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v:
            out[str(v).strip()] = c
    return out


# Map workbook header -> Client attribute.
HEADER_ATTR = {h: a for h, a, _w in COLUMNS}
HEADER_ATTR["Workflow Status"] = "status"


def _read_row(ws, r: int, hidx: dict) -> Client | None:
    def get(header):
        c = hidx.get(header)
        return ws.cell(r, c).value if c else None

    name = get("Name")
    app_id = get("App ID")
    if not name and not app_id:
        return None
    cl = Client(
        slide_index=int(get("Slide #") or 0),
        app_id=str(app_id or "").strip(),
        application_type=str(get("App Type") or "").strip(),
        name=str(name or "").strip(),
        application_date=str(get("Application Date") or "").strip(),
        age=str(get("Age") or "").strip(),
        occupation=str(get("Occupation") or "").strip(),
        company=str(get("Company") or "").strip(),
        membership_plan=str(get("Membership Plan") or "").strip(),
        rate_amount=_to_amount(get("Rate (QAR)")),
        tag=str(get("Tag") or "").strip(),
        cec=str(get("CEC") or "").strip(),
        mobile=str(get("Mobile") or "").strip(),
        special_request=str(get("Special Request / Note") or "").strip(),
        status=str(get("Workflow Status") or "Pending approval").strip(),
        photo=str(get("Photo file") or "").strip(),
    )
    cl.rate = "QAR %s" % format(cl.rate_amount, ",") if cl.rate_amount else ""
    if cl.status not in STATUSES:
        cl.status = "Pending approval"
    return cl


def _read_meta(wb) -> dict:
    meta = {}
    if "_meta" in wb.sheetnames:
        ws = wb["_meta"]
        for r in range(1, ws.max_row + 1):
            k, v = ws.cell(r, 1).value, ws.cell(r, 2).value
            if k:
                meta[str(k)] = v
    return meta


def sync_workbook(
    xlsx_path: str,
    media_dir: str | None,
    out_odp: str,
    out_xlsx: str | None,
    template_odp: str | None = None,
    out_pptx: str | None = None,
) -> int:
    if not os.path.exists(xlsx_path):
        print("ERROR: file not found: %s" % xlsx_path, file=sys.stderr)
        return 2
    wb = openpyxl.load_workbook(xlsx_path)
    meta = _read_meta(wb)
    template_odp = template_odp or meta.get("template_odp") or ""
    media_dir = media_dir or meta.get("media_dir") or "media"

    # Existing members from the three status sheets.
    existing: list[Client] = []
    for status in STATUSES:
        if status not in wb.sheetnames:
            continue
        ws = wb[status]
        hidx = _header_index(ws, 1)
        for r in range(2, ws.max_row + 1):
            cl = _read_row(ws, r, hidx)
            if cl:
                existing.append(cl)

    # New members from the Create Slide sheet (header on row 6).
    new_clients: list[Client] = []
    if "Create Slide" in wb.sheetnames:
        ws = wb["Create Slide"]
        hidx = _header_index(ws, 6)
        if hidx:
            for r in range(7, ws.max_row + 1):
                cl = _read_row(ws, r, hidx)
                if cl is None:
                    continue
                # "Create Slide" uses a "Photo file" column for the image path.
                pc = hidx.get("Photo file")
                if pc and ws.cell(r, pc).value:
                    cl.photo = str(ws.cell(r, pc).value).strip()
                new_clients.append(cl)

    # Assign slide indices / placeholder IDs to new members.
    next_idx = max([c.slide_index for c in existing] + [0]) + 1
    for i, cl in enumerate(new_clients):
        cl.slide_index = next_idx + i
        if not cl.app_id:
            cl.app_id = "MEM-APP-NEW-%03d" % (i + 1)

    print("Read %d existing + %d new member(s)" % (len(existing), len(new_clients)))
    from collections import Counter

    for s, n in Counter(c.status for c in existing + new_clients).most_common():
        print("    %-18s %d" % (s, n))

    # Regenerate the deck (design-faithful) if we have a template.
    if template_odp and os.path.exists(template_odp):
        from .slide_builder import build_deck

        build_deck(template_odp, existing, new_clients, out_odp)
        print("Rebuilt deck -> %s" % out_odp)
    else:
        print("WARNING: template deck not found (%r); skipping ODP rebuild. "
              "Pass --template <original.odp>." % template_odp, file=sys.stderr)

    all_clients = existing + new_clients

    # Native PowerPoint deck (sections + analysis), built from scratch.
    if out_pptx:
        from .pptx_builder import build_pptx

        build_pptx(all_clients, media_dir, out_pptx)
        print("Built PowerPoint -> %s" % out_pptx)

    # Rebuild the workbook with everyone re-filed by status.
    out_xlsx = out_xlsx or xlsx_path
    build_workbook(all_clients, media_dir, out_xlsx, template_path=template_odp)
    print("Rewrote workbook -> %s" % out_xlsx)
    return 0
