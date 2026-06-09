"""End-to-end tests for the Katara tracker: parse -> build -> sync round-trip
on a synthetic deck. Runnable with pytest or directly (``python test_tracker.py``).
"""

from __future__ import annotations

import os
import sys
import tempfile
import zipfile
from collections import Counter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

import openpyxl  # noqa: E402

from katara_tracker.odp_parser import parse_odp, extract_media  # noqa: E402
from katara_tracker.excel_builder import build_workbook  # noqa: E402
from katara_tracker.sync import sync_workbook  # noqa: E402
from katara_tracker.tests.fixtures import make_sample_odp  # noqa: E402

MEMBERS = [
    {"app_id": "MEM-APP-2025-00001", "name": "Alice Pending", "age": "30 yrs",
     "occupation": "Doctor", "company": "Hospital", "rate": "QAR 17,000",
     "membership_plan": "Single Membership - 6 Months", "cec": "Ines",
     "mobile": "30000001", "application_date": "01/01/2025", "tag": "DCL",
     "special_request": "wants discount", "status": "Pending approval",
     "photo": "media/p1.png"},
    {"app_id": "MEM-APP-2025-00002", "name": "Bob Approved", "age": "40 yrs",
     "occupation": "Lawyer", "company": "Firm", "rate": "QAR 19,000",
     "membership_plan": "Single Membership - 12 Months", "cec": "Asma",
     "mobile": "30000002", "application_date": "02/02/2025", "tag": "—",
     "special_request": "approved rate", "status": "Approved",
     "photo": "media/p2.png"},
    {"app_id": "MEM-APP-2025-00003", "name": "Carol Paid", "age": "25 yrs",
     "occupation": "Artist", "company": "Studio", "rate": "QAR 21,000",
     "membership_plan": "QF Single Membership - 12 Months", "cec": "Hilal",
     "mobile": "30000003", "application_date": "03/03/2025", "tag": "—",
     "special_request": "paid in full", "status": "Paid", "photo": "media/p3.png"},
]


def _setup(tmp):
    odp = os.path.join(tmp, "sample.odp")
    make_sample_odp(odp, MEMBERS)
    media = os.path.join(tmp, "media")
    extract_media(odp, media)
    return odp, media


def test_parse():
    with tempfile.TemporaryDirectory() as tmp:
        odp, _ = _setup(tmp)
        clients = parse_odp(odp)
        assert len(clients) == 3
        by_id = {c.app_id: c for c in clients}
        assert by_id["MEM-APP-2025-00001"].status == "Pending approval"
        assert by_id["MEM-APP-2025-00002"].status == "Approved"
        assert by_id["MEM-APP-2025-00003"].status == "Paid"
        c = by_id["MEM-APP-2025-00001"]
        assert c.name == "Alice Pending"
        assert c.rate_amount == 17000
        assert c.mobile == "30000001"
        assert c.occupation == "Doctor"
        assert c.application_date == "01/01/2025"


def test_build_workbook():
    with tempfile.TemporaryDirectory() as tmp:
        odp, media = _setup(tmp)
        clients = parse_odp(odp)
        out = os.path.join(tmp, "wb.xlsx")
        build_workbook(clients, media, out, template_path=odp)
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == [
            "Pending approval", "Approved", "Paid", "Analysis", "Create Slide",
            "_meta",
        ]
        assert wb["Pending approval"].max_row == 2  # header + 1
        assert wb["Paid"].max_row == 2
        # Analysis total paid == Carol's rate.
        wa = wb["Analysis"]
        vals = {wa.cell(r, 1).value: wa.cell(r, 2).value for r in range(1, 12)}
        assert vals["TOTAL AMOUNT PAID (all joiners)"] == 21000
        assert vals["Total joiners (clients)"] == 3


def test_sync_status_change_and_new_member():
    with tempfile.TemporaryDirectory() as tmp:
        odp, media = _setup(tmp)
        clients = parse_odp(odp)
        xlsx = os.path.join(tmp, "wb.xlsx")
        build_workbook(clients, media, xlsx, template_path=odp)

        # Move Alice (Pending) -> Paid, and add a new member via Create Slide.
        wb = openpyxl.load_workbook(xlsx)
        ws = wb["Pending approval"]
        h = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}
        ws.cell(2, h["Workflow Status"]).value = "Paid"
        cs = wb["Create Slide"]
        ch = {cs.cell(6, c).value: c for c in range(1, cs.max_column + 1)}
        photo = os.path.join(media, "image1.png")
        for k, v in {
            "Name": "Dan New", "Membership Plan": "Single Membership - 6 Months",
            "Rate (QAR)": 17000, "CEC": "Sara", "Workflow Status": "Approved",
            "Photo file": photo,
        }.items():
            cs.cell(7, ch[k]).value = v
        wb.save(xlsx)

        out_odp = os.path.join(tmp, "rebuilt.odp")
        rc = sync_workbook(xlsx, None, out_odp, None)
        assert rc == 0

        # Rebuilt deck is a valid ODF with 4 slides; statuses re-filed.
        assert zipfile.ZipFile(out_odp).testzip() is None
        reparsed = parse_odp(out_odp)
        assert len(reparsed) == 4
        counts = Counter(c.status for c in reparsed)
        assert counts["Paid"] == 2 and counts["Approved"] == 2
        assert counts["Pending approval"] == 0
        names = {c.name for c in reparsed}
        assert "Dan New" in names

        # Workbook re-filed: Alice now in Paid, Dan in Approved.
        wb2 = openpyxl.load_workbook(xlsx)
        paid_ids = [
            wb2["Paid"].cell(r, 3).value for r in range(2, wb2["Paid"].max_row + 1)
        ]
        assert "MEM-APP-2025-00001" in paid_ids


def test_sections_and_pptx():
    """The rebuilt decks include section dividers + analysis, and the native
    .pptx opens with the expected slide count."""
    import zipfile as _zip

    from lxml import etree

    from katara_tracker.pptx_builder import build_pptx

    with tempfile.TemporaryDirectory() as tmp:
        odp = os.path.join(tmp, "sample.odp")
        make_sample_odp(odp, MEMBERS)
        media = os.path.join(tmp, "media")
        extract_media(odp, media)
        xlsx = os.path.join(tmp, "wb.xlsx")
        build_workbook(parse_odp(odp), media, xlsx, template_path=odp)

        out_odp = os.path.join(tmp, "rebuilt.odp")
        out_pptx = os.path.join(tmp, "rebuilt.pptx")
        rc = sync_workbook(xlsx, media, out_odp, None, out_pptx=out_pptx)
        assert rc == 0

        # ODP: member slides skipped on reparse, but raw pages include sections.
        assert len(parse_odp(out_odp)) == 3
        root = etree.fromstring(_zip.ZipFile(out_odp).read("content.xml"))
        names = [
            p.get("{urn:oasis:names:tc:opendocument:xmlns:drawing:1.0}name")
            for p in root.iter(
                "{urn:oasis:names:tc:opendocument:xmlns:drawing:1.0}page")
        ]
        # 3 members + 3 status dividers + 1 analysis divider + 2 analysis pages.
        assert len(names) == 9
        assert names.count("Section") == 4

        # PPTX opens and has the same slide total.
        from pptx import Presentation

        prs = Presentation(out_pptx)
        assert len(list(prs.slides)) == 9


if __name__ == "__main__":
    test_parse()
    test_build_workbook()
    test_sync_status_change_and_new_member()
    test_sections_and_pptx()
    print("All tests passed.")
