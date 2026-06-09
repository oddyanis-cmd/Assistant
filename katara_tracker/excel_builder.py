"""Build the Katara member tracking workbook from parsed client records.

Layout (sheet order):
  1. Pending approval  - clients with no green note
  2. Approved          - clients with the green "Approved" note
  3. Paid              - clients with the green "Paid" note
  4. Analysis          - membership-type breakdown, totals, KPIs
  5. Create Slide      - a form to add new clients (+ photo) -> new slides on sync

Each client row carries a "Workflow Status" dropdown. Re-filing a client into
the matching sheet happens on ``sync`` (see cli.py), which also rebuilds the
deck so a status change moves the client's slide to match.
"""

from __future__ import annotations

import os
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .theme import GREEN, MAROON, MONEY_FMT, STATUS_COLORS, STATUSES, WHITE
from .odp_parser import Client

try:
    from PIL import Image as PILImage
except Exception:  # pragma: no cover
    PILImage = None

# Column spec: (header, Client attribute, width). "" attr = computed/blank.
COLUMNS = [
    ("Slide #", "slide_index", 8),
    ("Photo", "_photo_cell", 12),
    ("App ID", "app_id", 20),
    ("Workflow Status", "status", 18),
    ("Name", "name", 30),
    ("App Type", "application_type", 10),
    ("Application Date", "application_date", 16),
    ("Age", "age", 10),
    ("Occupation", "occupation", 24),
    ("Company", "company", 28),
    ("Membership Plan", "membership_plan", 28),
    ("Rate (QAR)", "rate_amount", 14),
    ("Tag", "tag", 8),
    ("CEC", "cec", 12),
    ("Mobile", "mobile", 14),
    ("Special Request / Note", "special_request", 50),
    ("Photo file", "photo", 22),
]

HEADER_FILL = PatternFill("solid", fgColor=MAROON)
HEADER_FONT = Font(bold=True, color=WHITE, size=11)
TITLE_FONT = Font(bold=True, color=MAROON, size=16)
LABEL_FONT = Font(bold=True, color=WHITE)
THIN = Side(style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
ROW_H = 50  # tall enough to show a thumbnail


def _thumb(path: str, max_h: int = 60):
    """Return an openpyxl image (thumbnail) for a media file, or None."""
    if not path or not os.path.exists(path) or PILImage is None:
        return None
    try:
        im = PILImage.open(path).convert("RGB")
        ratio = max_h / float(im.height)
        im = im.resize((max(1, int(im.width * ratio)), max_h))
        buf = BytesIO()
        im.save(buf, format="PNG")
        buf.seek(0)
        xi = XLImage(buf)
        xi.width, xi.height = im.width, im.height
        return xi
    except Exception:
        return None


def _style_header(ws, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = BORDER
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"


def _status_validation(ws, nrows: int, status_col: int) -> None:
    dv = DataValidation(
        type="list",
        formula1='"%s"' % ",".join(STATUSES),
        allow_blank=False,
        showDropDown=False,
    )
    dv.error = "Pick one of: " + ", ".join(STATUSES)
    dv.errorTitle = "Invalid workflow status"
    ws.add_data_validation(dv)
    col = get_column_letter(status_col)
    # Apply generously so newly pasted rows keep the dropdown.
    dv.add("%s2:%s%d" % (col, col, max(nrows + 1, 500)))


def _write_client_rows(ws, clients: list[Client], media_dir: str,
                       thumbnails: bool = True) -> None:
    """Write a header + one row per client onto an existing worksheet.

    ``thumbnails=False`` keeps the sheet as a clean data table (no floating
    images) so VBA can move rows between sheets safely.
    """
    for i, (header, _attr, width) in enumerate(COLUMNS, start=1):
        ws.cell(row=1, column=i, value=header)
        ws.column_dimensions[get_column_letter(i)].width = width
    _style_header(ws, len(COLUMNS))

    status_col = next(
        i for i, (_, a, _) in enumerate(COLUMNS, start=1) if a == "status"
    )
    rate_col = next(
        i for i, (_, a, _) in enumerate(COLUMNS, start=1) if a == "rate_amount"
    )
    photo_col = next(
        i for i, (_, a, _) in enumerate(COLUMNS, start=1) if a == "_photo_cell"
    )

    for r, client in enumerate(clients, start=2):
        ws.row_dimensions[r].height = ROW_H
        for c, (_header, attr, _w) in enumerate(COLUMNS, start=1):
            if attr in ("", "_photo_cell"):
                continue
            cell = ws.cell(row=r, column=c, value=getattr(client, attr, ""))
            cell.border = BORDER
            cell.alignment = WRAP if attr == "special_request" else Alignment(
                vertical="center", wrap_text=True
            )
        ws.cell(row=r, column=rate_col).number_format = MONEY_FMT
        ws.cell(row=r, column=status_col).alignment = CENTER
        # Embed the portrait thumbnail (look in media dir, then the raw path).
        if client.photo and thumbnails:
            candidates = [
                os.path.join(media_dir, os.path.basename(client.photo)),
                client.photo,
            ]
            img = next((t for t in (_thumb(p) for p in candidates) if t), None)
            if img is not None:
                ws.add_image(img, "%s%d" % (get_column_letter(photo_col), r))
        ws.cell(row=r, column=photo_col).border = BORDER

    _status_validation(ws, len(clients), status_col)


def _build_analysis(ws, clients: list[Client]) -> None:
    from collections import defaultdict

    ws.sheet_view.showGridLines = False
    ws["A1"] = "Katara Membership — Analysis"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    paid = [c for c in clients if c.status == "Paid"]
    approved = [c for c in clients if c.status == "Approved"]
    pending = [c for c in clients if c.status == "Pending approval"]
    total_paid = sum(c.rate_amount for c in paid)
    total_pipeline = sum(c.rate_amount for c in clients)

    kpis = [
        ("Total joiners (clients)", len(clients)),
        ("Pending approval", len(pending)),
        ("Approved (not paid)", len(approved)),
        ("Paid", len(paid)),
        ("TOTAL AMOUNT PAID (all joiners)", total_paid),
        ("Total pipeline value (all statuses)", total_pipeline),
        ("Average rate (all)", round(total_pipeline / len(clients)) if clients else 0),
    ]
    r = 3
    for label, value in kpis:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, color=MAROON)
        vc = ws.cell(row=r, column=2, value=value)
        if "AMOUNT" in label or "value" in label or "rate" in label.lower():
            vc.number_format = MONEY_FMT
            if "PAID" in label:
                vc.font = Font(bold=True, color=GREEN, size=12)
        r += 1

    # Membership-type breakdown.
    r += 1
    headers = ["Membership Type", "Joiners", "Paid", "Amount Paid (QAR)",
               "Pipeline (QAR)"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
    header_row = r
    r += 1

    by_type = defaultdict(lambda: {"n": 0, "paid": 0, "paid_amt": 0, "pipe": 0})
    for c in clients:
        key = c.membership_plan or "(unspecified)"
        d = by_type[key]
        d["n"] += 1
        d["pipe"] += c.rate_amount
        if c.status == "Paid":
            d["paid"] += 1
            d["paid_amt"] += c.rate_amount

    for key in sorted(by_type, key=lambda k: -by_type[k]["n"]):
        d = by_type[key]
        ws.cell(row=r, column=1, value=key)
        ws.cell(row=r, column=2, value=d["n"])
        ws.cell(row=r, column=3, value=d["paid"])
        ws.cell(row=r, column=4, value=d["paid_amt"]).number_format = MONEY_FMT
        ws.cell(row=r, column=5, value=d["pipe"]).number_format = MONEY_FMT
        r += 1
    # Totals row.
    ws.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=r, column=2, value="=SUM(B%d:B%d)" % (header_row + 1, r - 1))
    ws.cell(row=r, column=3, value="=SUM(C%d:C%d)" % (header_row + 1, r - 1))
    ws.cell(row=r, column=4,
            value="=SUM(D%d:D%d)" % (header_row + 1, r - 1)).number_format = MONEY_FMT
    ws.cell(row=r, column=5,
            value="=SUM(E%d:E%d)" % (header_row + 1, r - 1)).number_format = MONEY_FMT
    for c in range(1, 6):
        ws.cell(row=r, column=c).font = Font(bold=True)

    for col, w in zip("ABCDE", (34, 12, 10, 18, 16)):
        ws.column_dimensions[col].width = w


def _build_create_slide(ws) -> None:
    """A form/table for adding new clients that become slides on sync."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Create Slide — add new members here"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    notes = (
        "Fill ONE row per new member, then run:  python -m katara_tracker sync\n"
        "• A new slide is generated for each row and the member is filed into the "
        "sheet matching its Workflow Status.\n"
        "• Photo file: put the image path (e.g. media/john.jpg or an absolute "
        "path). It will appear on the generated slide."
    )
    ws["A2"] = notes
    ws.merge_cells("A2:H4")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 56

    # Header row identical to the status sheets, plus a Photo file column.
    headers = [h for h, _a, _w in COLUMNS if h != "Slide #"] + ["Photo file"]
    hr = 6
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
    widths = [w for h, _a, w in COLUMNS if h != "Slide #"] + [30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A%d" % (hr + 1)

    # Workflow Status dropdown on the new-member rows.
    status_idx = headers.index("Workflow Status") + 1
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(STATUSES),
                        allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    col = get_column_letter(status_idx)
    dv.add("%s%d:%s%d" % (col, hr + 1, col, hr + 200))
    # A few pre-formatted blank rows for convenience.
    for r in range(hr + 1, hr + 16):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = BORDER


# --- Membership-card gallery -------------------------------------------------
# A card occupies CARD_W columns x CARD_H rows; CARDS_PER_ROW cards per band,
# separated by one gutter column/row.
CARD_W = 7
CARD_H = 9
CARDS_PER_ROW = 3
GUTTER = 1
CARD_BORDER = Border(*(Side(style="thin", color=MAROON),) * 4)
CARD_TOP = Side(style="medium", color=MAROON)


def _draw_card(ws, client: Client, top: int, left: int, media_dir: str) -> None:
    right = left + CARD_W - 1
    status_color = STATUS_COLORS.get(client.status, MAROON)

    # Header band: title + coloured status badge.
    ws.merge_cells(start_row=top, start_column=left, end_row=top,
                   end_column=left + 4)
    h = ws.cell(top, left, "MEMBER PROFILE")
    h.fill = HEADER_FILL
    h.font = Font(bold=True, color=WHITE, size=10)
    h.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=top, start_column=left + 5, end_row=top,
                   end_column=right)
    badge_text = "Pending" if client.status == "Pending approval" else client.status
    b = ws.cell(top, left + 5, badge_text)
    b.fill = PatternFill("solid", fgColor=status_color)
    b.font = Font(bold=True, color=WHITE, size=9)
    b.alignment = CENTER
    ws.row_dimensions[top].height = 20

    # Portrait spanning the left of the card body.
    if client.photo:
        img = _thumb(os.path.join(media_dir, os.path.basename(client.photo)),
                     max_h=92)
        if img is not None:
            ws.add_image(img, "%s%d" % (get_column_letter(left), top + 1))

    # Info column to the right of the photo.
    info = left + 2
    maroon_bold = Font(bold=True, color=MAROON, size=11)
    grey = Font(color="777777", size=8)

    def line(roff, value, font=None, money=False):
        r = top + roff
        ws.merge_cells(start_row=r, start_column=info, end_row=r,
                       end_column=right)
        cell = ws.cell(r, info, value)
        cell.font = font or Font(size=9)
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        if money:
            cell.number_format = MONEY_FMT
        ws.row_dimensions[r].height = 16
        return cell

    line(1, client.name, maroon_bold)
    line(2, client.app_id, grey)
    line(3, client.membership_plan, Font(size=9))
    line(4, client.rate_amount, Font(bold=True, color=status_color, size=10),
         money=True)
    line(5, "CEC: %s" % (client.cec or "—"), Font(size=9))
    line(6, "Mob: %s" % (client.mobile or "—"), Font(size=9))
    line(7, "Applied: %s" % (client.application_date or "—"), Font(size=8,
         color="555555"))

    # Outline the whole card.
    for r in range(top, top + CARD_H - 1):
        for c in range(left, right + 1):
            cell = ws.cell(r, c)
            edges = {}
            if r == top:
                edges["top"] = CARD_TOP
            if r == top + CARD_H - 2:
                edges["bottom"] = CARD_BORDER.bottom
            if c == left:
                edges["left"] = CARD_BORDER.left
            if c == right:
                edges["right"] = CARD_BORDER.right
            if edges:
                cell.border = Border(**edges)


def _build_cards(ws, clients: list[Client], media_dir: str) -> None:
    """Lay members out as a gallery of membership cards (grouped by status)."""
    ws.sheet_view.showGridLines = False
    ws["A1"] = "Member Cards — visual gallery"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")

    # Uniform card column widths (+ narrow gutter columns).
    for g in range(CARDS_PER_ROW):
        base = 1 + g * (CARD_W + GUTTER)
        for c in range(base, base + CARD_W):
            ws.column_dimensions[get_column_letter(c)].width = 11
        if g < CARDS_PER_ROW - 1:
            ws.column_dimensions[get_column_letter(base + CARD_W)].width = 2

    order = {"Pending approval": 0, "Approved": 1, "Paid": 2}
    ordered = sorted(clients, key=lambda c: (order.get(c.status, 0),
                                             c.slide_index))
    start_row = 3
    for k, client in enumerate(ordered):
        gr, gc = divmod(k, CARDS_PER_ROW)
        left = 1 + gc * (CARD_W + GUTTER)
        top = start_row + gr * (CARD_H + GUTTER)
        _draw_card(ws, client, top, left, media_dir)


def build_workbook(
    clients: list[Client],
    media_dir: str,
    out_path: str,
    template_path: str = "",
    macro_ready: bool = False,
) -> None:
    """Build the tracking workbook.

    ``macro_ready=True`` keeps the status sheets as clean data tables (no
    floating thumbnails) so the live VBA automation can move rows between
    sheets safely; the Member Cards sheet still carries the photos.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for status in STATUSES:
        ws = wb.create_sheet(title=status)
        subset = [c for c in clients if c.status == status]
        _write_client_rows(ws, subset, media_dir, thumbnails=not macro_ready)

    _build_cards(wb.create_sheet(title="Member Cards"), clients, media_dir)
    _build_analysis(wb.create_sheet(title="Analysis"), clients)
    _build_create_slide(wb.create_sheet(title="Create Slide"))

    # Hidden bookkeeping sheet so `sync` knows the template deck + media dir.
    meta = wb.create_sheet(title="_meta")
    meta["A1"] = "template_odp"
    meta["B1"] = os.path.abspath(template_path) if template_path else ""
    meta["A2"] = "media_dir"
    meta["B2"] = os.path.abspath(media_dir) if media_dir else ""
    meta["A3"] = "macro_ready"
    meta["B3"] = "1" if macro_ready else ""
    meta.sheet_state = "hidden"

    wb.save(out_path)
