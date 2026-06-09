"""Read/write a flat member table (CSV or single-sheet Excel).

This is the bridge between the no-code app (Microsoft Lists / Google Sheets /
Airtable) and the deck generator:

* ``write_import_table`` produces a clean one-table workbook to *import* into
  Microsoft Lists (it becomes the list's columns + rows).
* ``read_table`` reads a list *export* (CSV or Excel) back into Client records
  so the PowerPoint/ODP decks and the tracking workbook can be regenerated.
"""

from __future__ import annotations

import csv
import os

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

from .odp_parser import Client
from .theme import STATUSES

# (column header shown to staff, Client attribute)
TABLE_COLUMNS = [
    ("Name", "name"),
    ("App ID", "app_id"),
    ("Workflow Status", "status"),
    ("Application Type", "application_type"),
    ("Application Date", "application_date"),
    ("Age", "age"),
    ("Occupation", "occupation"),
    ("Company", "company"),
    ("Membership Plan", "membership_plan"),
    ("Rate (QAR)", "rate_amount"),
    ("Tag", "tag"),
    ("CEC", "cec"),
    ("Mobile", "mobile"),
    ("Special Request", "special_request"),
    ("Photo file", "photo"),
]
HEADERS = [h for h, _a in TABLE_COLUMNS]


def write_import_table(clients: list[Client], out_path: str) -> None:
    """Write a single-table .xlsx ready for 'Create a list from Excel'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    ws.append(HEADERS)
    for c in clients:
        ws.append([_value(c, attr) for _h, attr in TABLE_COLUMNS])

    # Define an Excel Table so Microsoft Lists detects the columns cleanly.
    last_col = chr(ord("A") + len(HEADERS) - 1)
    ref = "A1:%s%d" % (last_col, len(clients) + 1)
    table = Table(displayName="Members", ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showRowStripes=True
    )
    ws.add_table(table)
    for i, h in enumerate(HEADERS):
        ws.column_dimensions[chr(ord("A") + i)].width = max(12, len(h) + 2)
    wb.save(out_path)


def write_import_csv(clients: list[Client], out_path: str) -> None:
    with open(out_path, "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(HEADERS)
        for c in clients:
            w.writerow([_value(c, attr) for _h, attr in TABLE_COLUMNS])


def _value(c: Client, attr: str):
    v = getattr(c, attr, "")
    return v if v not in (None,) else ""


def read_table(path: str) -> list[Client]:
    """Read a CSV or single-sheet Excel export into Client records."""
    rows = _read_rows(path)
    if not rows:
        return []
    header = [str(h or "").strip() for h in rows[0]]
    idx = {h.lower(): i for i, h in enumerate(header)}

    clients = []
    for n, raw in enumerate(rows[1:], start=1):
        get = lambda key: _cell(raw, idx, key)
        name = get("name")
        app_id = get("app id")
        if not name and not app_id:
            continue
        cl = Client(
            slide_index=n,
            name=name,
            app_id=app_id,
            status=_norm_status(get("workflow status")),
            application_type=get("application type"),
            application_date=get("application date"),
            age=get("age"),
            occupation=get("occupation"),
            company=get("company"),
            membership_plan=get("membership plan"),
            rate_amount=_to_int(get("rate (qar)") or get("rate")),
            tag=get("tag"),
            cec=get("cec"),
            mobile=get("mobile"),
            special_request=get("special request") or get("special note"),
            photo=get("photo file") or get("photo"),
        )
        cl.rate = "QAR %s" % format(cl.rate_amount, ",") if cl.rate_amount else ""
        if not cl.app_id:
            cl.app_id = "MEM-APP-NEW-%03d" % n
        clients.append(cl)
    return clients


def _read_rows(path: str) -> list[list]:
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as fh:
            return [list(r) for r in csv.reader(fh)]
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    return [[c.value for c in row] for row in ws.iter_rows()]


def _cell(raw, idx, key) -> str:
    i = idx.get(key)
    if i is None or i >= len(raw):
        return ""
    v = raw[i]
    return "" if v is None else str(v).strip()


def _to_int(v) -> int:
    if not v:
        return 0
    s = "".join(ch for ch in str(v) if ch.isdigit())
    return int(s) if s else 0


def _norm_status(v: str) -> str:
    s = (v or "").strip().lower()
    if s in ("paid",):
        return "Paid"
    if s in ("approved",):
        return "Approved"
    if s in ("pending approval", "pending", ""):
        return "Pending approval"
    # Match any provided status case-insensitively, else default.
    for st in STATUSES:
        if st.lower() == s:
            return st
    return "Pending approval"
